import xml.etree.ElementTree as ET
import sqlite3
import sys
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Font 
import time
import datetime as dt


rutaArchivos=sys.argv[1]

def generaXls(db,id):
    connection = sqlite3.connect(db)
   
    l1=["Tipo Doc","Clave Acceso", "RUC Proveedor","Nombre Proveedor","RUC Cliente",
        "Nombre Cliente","Establecimeinto","Pto Emisión","Secuencial",
        "Base No O","Base 0","Base I","Base Ex","ICE","IVA","Total"] # List of column header 

    l2=["Tipo Doc","Clave Acceso", "RUC Proveedor","Nombre Proveedor","RUC Cliente",
        "Nombre Cliente","Establecimeinto","Pto Emisión","Secuencial","Doc Sustento",
        "Fecha Doc","Impuesto","Cod.","Cod. Ret","Base I","% Ret.","Valor"]
    #query="SELECT tipoDocumento,claveAcceso,rucProveedor,razonSocialP,rucCliente,nombreCliente,establecimiento,puntoEmision,secuencial,totalSinImp,valorTotal FROM documentos" # query 
    # CAST(a.totalSinImp AS decimal),
    query="""
    SELECT a.tipoDocumento,a.claveAcceso,a.rucProveedor,a.razonSocialP,a.rucCliente,
    a.nombreCliente,a.establecimiento,a.puntoEmision,a.secuencial,
    CAST((select sum(baseImponible) from detalleRetFactura where idComprobante=a.id and codigo=2 and codigoPorcentaje=6) AS decimal) as baseNo,
    CAST((select sum(baseImponible) from detalleRetFactura where idComprobante=a.id and codigo=2 and codigoPorcentaje=0) AS decimal) as base0,
    CAST((select sum(baseImponible) from detalleRetFactura where idComprobante=a.id and codigo=2 and codigoPorcentaje=2) AS decimal) as base12,
    CAST((select sum(baseImponible) from detalleRetFactura where idComprobante=a.id and codigo=2 and codigoPorcentaje=7) AS decimal) as baseE,
    CAST((select sum(valor) from detalleRetFactura where idComprobante=a.id and codigo=3) AS decimal) as montoIce,
    CAST((select sum(valor) from detalleRetFactura where idComprobante=a.id and codigo=2) AS decimal) as montoIva,
    CAST((select sum(baseImponible) from detalleRetFactura where idComprobante=a.id) AS decimal) as bases

    FROM ejecucionDocumentos as b INNER JOIN
    documentos as a ON b.idDocumento=a.id
    WHERE a.tipoDocumento='FACTURA' and b.idEjecucion=?

    """
    
    query3="""
    SELECT a.tipoDocumento,a.claveAcceso,a.rucProveedor,a.razonSocialP,a.rucCliente,
    a.nombreCliente,a.establecimiento,a.puntoEmision,a.secuencial,
    CAST((select sum(baseImponible) from detalleNotaCredito where idComprobante=a.id and codigo=2 and codigoPorcentaje=6) AS decimal) as baseNo,
    CAST((select sum(baseImponible) from detalleNotaCredito where idComprobante=a.id and codigo=2 and codigoPorcentaje=0) AS decimal) as base0,
    CAST((select sum(baseImponible) from detalleNotaCredito where idComprobante=a.id and codigo=2 and codigoPorcentaje=2) AS decimal) as base12,
    CAST((select sum(baseImponible) from detalleNotaCredito where idComprobante=a.id and codigo=2 and codigoPorcentaje=7) AS decimal) as baseE,
    CAST((select sum(valor) from detalleNotaCredito where idComprobante=a.id and codigo=3) AS decimal) as montoIce,
    CAST((select sum(valor) from detalleNotaCredito where idComprobante=a.id and codigo=2) AS decimal) as montoIva,
    CAST((select sum(baseImponible) from detalleNotaCredito where idComprobante=a.id) AS decimal) as bases

    FROM ejecucionDocumentos as b INNER JOIN
    documentos as a ON b.idDocumento=a.id
    WHERE a.tipoDocumento LIKE 'N. CR%' and b.idEjecucion=?
    """
    
    query4="""
    SELECT a.tipoDocumento,a.claveAcceso,a.rucProveedor,a.razonSocialP,a.rucCliente,
    a.nombreCliente,a.establecimiento,a.puntoEmision,a.secuencial,
    CAST((select sum(baseImponible) from detalleNotaDebito where idComprobante=a.id and codigo=2 and codigoPorcentaje=6) AS decimal) as baseNo,
    CAST((select sum(baseImponible) from detalleNotaDebito where idComprobante=a.id and codigo=2 and codigoPorcentaje=0) AS decimal) as base0,
    CAST((select sum(baseImponible) from detalleNotaDebito where idComprobante=a.id and codigo=2 and codigoPorcentaje=2) AS decimal) as base12,
    CAST((select sum(baseImponible) from detalleNotaDebito where idComprobante=a.id and codigo=2 and codigoPorcentaje=7) AS decimal) as baseE,
    CAST((select sum(valor) from detalleNotaDebito where idComprobante=a.id and codigo=3) AS decimal) as montoIce,
    CAST((select sum(valor) from detalleNotaDebito where idComprobante=a.id and codigo=2) AS decimal) as montoIva,
    CAST((select sum(baseImponible) from detalleNotaDebito where idComprobante=a.id) AS decimal) as bases

    FROM ejecucionDocumentos as b INNER JOIN
    documentos as a ON b.idDocumento=a.id
    WHERE a.tipoDocumento LIKE 'N. D%' and b.idEjecucion=?
    """
    
    query2="""
     select a.tipoDocumento,a.claveAcceso,a.rucProveedor,a.razonSocialP,a.rucCliente,
    a.nombreCliente,a.establecimiento,a.puntoEmision,a.secuencial,b.docSustento,b.fechaDoc,
    b.impuesto,b.codigo,b.codigoRetencion,b.baseImponible,b.porcentajeRetenido,
    b.valorRetenido
    FROM ejecucionDocumentos as c INNER JOIN
    documentos as a ON c.idDocumento=a.id  INNER JOIN
    detalleRetencion as b ON b.idComprobante=a.id
    WHERE c.idEjecucion=?
    """
    my_data=connection.execute(query,(id,)) 
    my_data=[r for r in my_data] # List of rows of data 

    my_data2=connection.execute(query2,(id,)) 
    my_data2=[r for r in my_data2] # List of rows of data 

    my_data3=connection.execute(query3,(id,)) 
    my_data3=[r for r in my_data3] # List of rows of data 

    my_data4=connection.execute(query4,(id,)) 
    my_data4=[r for r in my_data4] # List of rows of data 

    wb=Workbook()
    ws1 = wb.create_sheet("Facturas_Notas")
    ws2 = wb.create_sheet("Retenciones")
    #ws1=wb.active # work with default worksheet
    ws1.append(l1) # adding column headers at first row 
    ws2.append(l2) 
    my_font=Font(size=14,bold=True) # font styles
    my_fill=PatternFill(fill_type='solid',start_color='267007') #Background color
    for cell in ws1["1:1"]: # First row 
        cell.font = my_font
        cell.fill= my_fill

    for cell in ws2["1:1"]: # First row 
        cell.font = my_font
        cell.fill= my_fill

    r,c=2,0 # row=2 and column=0
    for row_data in my_data:
        d=[r for r in row_data]
        ws1.append(d)

    for row_data in my_data3:
        d=[r for r in row_data]
        ws1.append(d)
    for row_data in my_data4:
        d=[r for r in row_data]
        ws1.append(d)
    r,c=2,0 # row=2 and column=0
    for row_data in my_data2:
        d=[r for r in row_data]
        ws2.append(d)
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 55
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 40
    ws1.column_dimensions['E'].width = 20
    ws1.column_dimensions['F'].width = 40
    ws1.column_dimensions['G'].width = 10
    ws1.column_dimensions['H'].width = 10
    ws1.column_dimensions['I'].width = 15
    ws1.column_dimensions['J'].width = 10
    ws1.column_dimensions['K'].width = 10
    ws1.column_dimensions['L'].width = 10
    ws1.column_dimensions['M'].width = 10
    ws1.column_dimensions['N'].width = 10
    ws1.column_dimensions['O'].width = 10
    ws1.column_dimensions['P'].width = 10
    #current_date = dt.date.today()

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 55
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 40
    ws2.column_dimensions['E'].width = 20
    ws2.column_dimensions['F'].width = 40
    ws2.column_dimensions['G'].width = 10
    ws2.column_dimensions['H'].width = 10
    ws2.column_dimensions['I'].width = 15
    ws2.column_dimensions['J'].width = 15
    ws2.column_dimensions['K'].width = 16
    ws2.column_dimensions['L'].width = 10
    ws2.column_dimensions['M'].width = 10
    ws2.column_dimensions['N'].width = 10
    ws2.column_dimensions['O'].width = 10
    ws2.column_dimensions['P'].width = 10
    ws2.column_dimensions['Q'].width = 10
    wb.remove(wb['Sheet'])
    # my_path=ruc+'.xlsx'#Path 
    current_date = dt.datetime.now()
    fechaActual=current_date.strftime("%Y%m%d%H%M%S")#Path 

    my_path=fechaActual+'.xlsx'#Path 
    wb.save(my_path)
    time.sleep(2)
    os.startfile(my_path)
def leerXML(path,db,idEjecucion):
    print(path)
    conexion=sqlite3.connect(db)
    #cur = conexion.cursor()
    root = ET.parse(path)
   
    root_node = root.getroot()
    estado=root_node.find("estado").text
    numeroAutorizacion=root_node.find("numeroAutorizacion").text
    fechaAutorizacion=root_node.find("fechaAutorizacion").text
    ambiente=root_node.find("ambiente").text
    comprobante=root_node.find("comprobante").text
    root_comprobante = ET.fromstring (comprobante.replace('<?xml version="1.0" encoding="UTF-8"?>',''))
    tipoComprobante=root_comprobante.tag
    version=root_comprobante.attrib["version"]
    print(tipoComprobante)
    if(tipoComprobante=="factura"):
        tipoDoc="FACTURA"
    if(tipoComprobante=="notaCredito"):
        tipoDoc="N. CRÉDITO"
    if(tipoComprobante=="notaDebito"):
        tipoDoc="N. DÉBITO"
    if(tipoComprobante=="comprobanteRetencion"):
        tipoDoc="C. RETENCIÓN"

    infoTributaria=root_comprobante.find("infoTributaria")
    claveAcceso=infoTributaria.find("claveAcceso").text
    contCursor = conexion.cursor()
    idDocumento=0
    cnt = contCursor.execute("select count(*) as contador from documentos where observacion='OK'and claveAcceso=?",(claveAcceso,))
    contador = cnt.fetchone()
    cur = conexion.cursor()
    if contador[0]==0:
        try:
            insertDoc=cur.execute("insert into documentos(claveAcceso,tipoDocumento,observacion) values (?,?,?)", 
            (str(claveAcceso),tipoDoc,'NO PROCESADO'))
            conexion.commit()
            idDocumento=insertDoc.lastrowid
        except:
            new_cur = conexion.cursor()
            res = new_cur.execute("SELECT observacion,id as idDoc FROM documentos WHERE claveAcceso=? ORDER BY id DESC",(claveAcceso,))
            observacion,idDoc = res.fetchone()
            idDocumento=idDoc
            if observacion=="NO PROCESADO":
                try:
                    new_cur.execute("DELETE FROM detalleFactura WHERE idComprobante=?",(idDocumento,))
                    conexion.commit()
                except:
                    pass
                try:
                    new_cur.execute("DELETE FROM detalleRetFactura WHERE idComprobante=?",(idDocumento,))
                    conexion.commit()
                except:
                    pass
                try:
                    new_cur.execute("DELETE FROM detalleNotaCredito WHERE idComprobante=?",(idDocumento,))
                    conexion.commit()
                except:
                    pass
                try:
                    new_cur.execute("DELETE FROM detalleNotaDebito WHERE idComprobante=?",(idDocumento,))
                    conexion.commit()
                except:
                    pass
                try:
                    new_cur.execute("DELETE FROM detalleRetencion WHERE idComprobante=?",(idDocumento,))
                    conexion.commit()
                except:
                    pass
    
    
    else:
        idDocumento=0
        new_cur = conexion.cursor()
        res = new_cur.execute("SELECT observacion,id as idDoc FROM documentos WHERE claveAcceso=? ORDER BY id DESC",(claveAcceso,))
        observacion,idDoc = res.fetchone()
        idInsertar=idDoc
        insertDoc=cur.execute("insert into ejecucionDocumentos(idDocumento,idEjecucion) values (?,?)", 
        (idInsertar,idEjecucion))
        conexion.commit()



    if(idDocumento>0):
       
        insertDoc=cur.execute("insert into ejecucionDocumentos(idDocumento,idEjecucion) values (?,?)", 
        (idDocumento,idEjecucion))
        conexion.commit()

        if(tipoDoc=="C. RETENCIÓN"):
            if version=="1.0.0":
            
                documentosSustento=root_comprobante.find("impuestos")
                ambiente=infoTributaria.find("ambiente").text
                tipoEmision=infoTributaria.find("tipoEmision").text
                razonSocial=infoTributaria.find("razonSocial").text
                nombreComercial=""
                try:
                    nombreComercial=infoTributaria.find("nombreComercial").text
                except:
                    pass
                #nombreComercial=infoTributaria.find("nombreComercial").text
                ruc=infoTributaria.find("ruc").text
                #claveAcceso=infoTributaria.find("claveAcceso").text
                codDoc=infoTributaria.find("codDoc").text
                estab=infoTributaria.find("estab").text
                ptoEmi=infoTributaria.find("ptoEmi").text
                secuencial=infoTributaria.find("secuencial").text
                dirMatriz=infoTributaria.find("dirMatriz").text

                infoCompRetencion=root_comprobante.find("infoCompRetencion")
                fechaEmision=infoCompRetencion.find("fechaEmision").text
                dirEstablecimiento=""
                try:

                    dirEstablecimiento=infoCompRetencion.find("dirEstablecimiento").text
                except:
                    pass
                
                contribuyenteEspecial=""
                obligadoContabilidad=""
                try:
                    contribuyenteEspecial=infoCompRetencion.find("contribuyenteEspecial").text
                except:
                    pass
                try:
                    obligadoContabilidad=infoCompRetencion.find("obligadoContabilidad").text
                except:
                    pass
                
            
                tipoIdentificacionSujetoRetenido=infoCompRetencion.find("tipoIdentificacionSujetoRetenido").text
                #parteRel=infoCompRetencion.find("parteRel").text
                razonSocialSujetoRetenido=infoCompRetencion.find("razonSocialSujetoRetenido").text
                identificacionSujetoRetenido=infoCompRetencion.find("identificacionSujetoRetenido").text
                periodoFiscal=""
                try:
                    periodoFiscal=infoCompRetencion.find("periodoFiscal").text
                except:
                    pass
                
                tipoProveedor=""
                if(contribuyenteEspecial==""):
                    tipoProveedor="01"
                else:
                    tipoProveedor="02"
                

                idSociedadInsertar=1

                cur.execute("update documentos set tipoDocumento='"+tipoDoc+"', tipoProveedor='"+tipoProveedor+"', tipoEmision='"+ tipoEmision+"',ambiente='"+ambiente+"',rucProveedor='"+ruc+"',nombreProveedor='"+nombreComercial+"',direccionMatriz='"+dirMatriz+"',direccionEstablecimiento='"+dirEstablecimiento+"',razonSocialP='"+razonSocial+"',tipoIdentificacionC='"+tipoIdentificacionSujetoRetenido+"',rucCliente='"+identificacionSujetoRetenido+"',nombreCliente='"+razonSocialSujetoRetenido+"',establecimiento='"+estab+"',puntoEmision='"+ptoEmi+"',secuencial='"+secuencial+"',fecha='"+fechaEmision+"',totalSinImp='',descuento='',propina='',devolucionIva='',valorTotal='',contrEspecial='"+contribuyenteEspecial+"',obligado='"+obligadoContabilidad+"',microempresa='',rimpe='',agenteRetencion='',ejercicioFiscal='"+periodoFiscal+"',idSociedad="+str(idSociedadInsertar)+" WHERE claveAcceso=?", 
                (claveAcceso,))
                conexion.commit()
                #detalles

                for child in documentosSustento.findall("impuesto"):
                    numDocSustento=''
                    fechaEmisionDocSustento=''
                    codigo=''
                    try:
                        numDocSustento=child.find("numDocSustento").text
                    except:
                        pass
                    try:
                        fechaEmisionDocSustento=child.find("fechaEmisionDocSustento").text
                    except:
                        pass
                    try:
                        codigo=child.find("codigo").text
                    except:
                        pass
                    
                    impuesto=""
                    if(str(codigo)=="2"):
                        impuesto="IVA"
                    if(str(codigo)=="1"):
                        impuesto="RENTA"
                    if(str(codigo)=="6"):
                        impuesto="ISD"

                    codigoRetencion=child.find("codigoRetencion").text
                    baseImponible=child.find("baseImponible").text
                    porcentajeRetener=child.find("porcentajeRetener").text
                    valorRetenido=child.find("valorRetenido").text
                    conexion.execute("insert into detalleRetencion (impuesto,codigo,codigoRetencion,baseImponible,porcentajeRetenido,valorRetenido,docSustento,fechaDoc,idComprobante) values (?,?,?,?,?,?,?,?,?)",
                    (impuesto,codigo,codigoRetencion,baseImponible,porcentajeRetener,valorRetenido,numDocSustento,fechaEmisionDocSustento,idDocumento))
                    conexion.commit()   
            else:
                documentosSustento=root_comprobante.find("docsSustento")
                #infoTributaria=root_comprobante.find("infoTributaria")
                ambiente=infoTributaria.find("ambiente").text
                tipoEmision=infoTributaria.find("tipoEmision").text
                razonSocial=infoTributaria.find("razonSocial").text
                nombreComercial=""
                try:
                    nombreComercial=infoTributaria.find("nombreComercial").text
                except:
                    pass
                #nombreComercial=infoTributaria.find("nombreComercial").text
                ruc=infoTributaria.find("ruc").text
                #claveAcceso=infoTributaria.find("claveAcceso").text
                codDoc=infoTributaria.find("codDoc").text
                estab=infoTributaria.find("estab").text
                ptoEmi=infoTributaria.find("ptoEmi").text
                secuencial=infoTributaria.find("secuencial").text
                dirMatriz=infoTributaria.find("dirMatriz").text

                infoCompRetencion=root_comprobante.find("infoCompRetencion")
                fechaEmision=infoCompRetencion.find("fechaEmision").text
                dirEstablecimiento=""
                try:

                    dirEstablecimiento=infoCompRetencion.find("dirEstablecimiento").text
                except:
                    pass
                
                contribuyenteEspecial=""
                obligadoContabilidad=""
                try:
                    contribuyenteEspecial=infoCompRetencion.find("contribuyenteEspecial").text
                except:
                    pass
                try:
                    obligadoContabilidad=infoCompRetencion.find("obligadoContabilidad").text
                except:
                    pass
                
            
                tipoIdentificacionSujetoRetenido=infoCompRetencion.find("tipoIdentificacionSujetoRetenido").text
                #parteRel=infoCompRetencion.find("parteRel").text
                razonSocialSujetoRetenido=infoCompRetencion.find("razonSocialSujetoRetenido").text
                identificacionSujetoRetenido=infoCompRetencion.find("identificacionSujetoRetenido").text
                periodoFiscal=""
                try:
                    periodoFiscal=infoCompRetencion.find("periodoFiscal").text
                except:
                    pass
                
                tipoProveedor=""
                if(contribuyenteEspecial==""):
                    tipoProveedor="01"
                else:
                    tipoProveedor="02"
                
                idSociedadInsertar=1
                
                cur.execute("update documentos set tipoDocumento='"+tipoDoc+"', tipoProveedor='"+tipoProveedor+"', tipoEmision='"+ tipoEmision+"',ambiente='"+ambiente+"',rucProveedor='"+ruc+"',nombreProveedor='"+nombreComercial+"',direccionMatriz='"+dirMatriz+"',direccionEstablecimiento='"+dirEstablecimiento+"',razonSocialP='"+razonSocial+"',tipoIdentificacionC='"+tipoIdentificacionSujetoRetenido+"',rucCliente='"+identificacionSujetoRetenido+"',nombreCliente='"+razonSocialSujetoRetenido+"',establecimiento='"+estab+"',puntoEmision='"+ptoEmi+"',secuencial='"+secuencial+"',fecha='"+fechaEmision+"',totalSinImp='',descuento='',propina='',devolucionIva='',valorTotal='',contrEspecial='"+contribuyenteEspecial+"',obligado='"+obligadoContabilidad+"',microempresa='',rimpe='',agenteRetencion='',ejercicioFiscal='"+periodoFiscal+"' ,idSociedad="+str(idSociedadInsertar)+" WHERE claveAcceso=?", 
                (claveAcceso,))
                conexion.commit()

                for child in documentosSustento.findall("docSustento"):
                    codSustento=child.find("codSustento").text
                    codDocSustento=child.find("codDocSustento").text
                    numDocSustento=child.find("numDocSustento").text
                    fechaEmisionDocSustento=child.find("fechaEmisionDocSustento").text
                    #fechaRegistroContable=child.find("fechaRegistroContable").text
                    #numAutDocSustento=child.find("numAutDocSustento").text
                    #pagoLocExt=child.find("pagoLocExt").text
                    #totalSinImpuestos=child.find("totalSinImpuestos").text
                    #importeTotal=child.find("importeTotal").text
                # print("Hijo de root_node: ", child.tag, child.attrib, child.text, child.tail)
                    
                    nodoRetenciones=child.find("retenciones")
                    for child2 in nodoRetenciones:
                        codigo=child2.find("codigo").text
                        impuesto=""
                        if(str(codigo)=="2"):
                            impuesto="IVA"
                        if(str(codigo)=="1"):
                            impuesto="RENTA"
                        if(str(codigo)=="6"):
                            impuesto="ISD"

                        codigoRetencion=child2.find("codigoRetencion").text
                        baseImponible=child2.find("baseImponible").text
                        porcentajeRetener=child2.find("porcentajeRetener").text
                        valorRetenido=child2.find("valorRetenido").text
                        #print(impuesto,codigo,codigoRetencion,baseImponible,porcentajeRetener,valorRetenido,numDocSustento,fechaEmisionDocSustento,idDoc)
                        conexion.execute("insert into detalleRetencion (impuesto,codigo,codigoRetencion,baseImponible,porcentajeRetenido,valorRetenido,docSustento,fechaDoc,idComprobante) values (?,?,?,?,?,?,?,?,?)",
                        (impuesto,codigo,codigoRetencion,baseImponible,porcentajeRetener,valorRetenido,numDocSustento,fechaEmisionDocSustento,idDocumento))
                        conexion.commit()
                        #print("codigo: "+codigo,"Cod Ret: "+codigoRetencion,"BI: "+baseImponible,"% "+porcentajeRetener,"valor: "+valorRetenido)

            cur.execute("UPDATE documentos set observacion='OK' WHERE id=?",(idDocumento,))
            conexion.commit()
        if(tipoDoc=="FACTURA"):
            infoFactura=root_comprobante.find("infoFactura")
            #infoTributaria=root_comprobante.find("infoTributaria")
            impuestos=infoFactura.find("totalConImpuestos")
            ambiente=infoTributaria.find("ambiente").text
            tipoEmision=infoTributaria.find("tipoEmision").text
            razonSocial=infoTributaria.find("razonSocial").text
            nombreComercial=""
            try:
                nombreComercial=infoTributaria.find("nombreComercial").text
            except:
                pass
            ruc=infoTributaria.find("ruc").text
            #claveAcceso=infoTributaria.find("claveAcceso").text
            codDoc=infoTributaria.find("codDoc").text
            estab=infoTributaria.find("estab").text
            ptoEmi=infoTributaria.find("ptoEmi").text
            secuencial=infoTributaria.find("secuencial").text
            dirMatriz=infoTributaria.find("dirMatriz").text

            
            fechaEmision=infoFactura.find("fechaEmision").text
            dirEstablecimiento=""
            try:
                dirEstablecimiento=infoFactura.find("dirEstablecimiento").text
            except:
                pass
                
            contribuyenteEspecial=""
            obligadoContabilidad=""
            try:
                contribuyenteEspecial=infoFactura.find("contribuyenteEspecial").text
            except:
                pass
            try:
                obligadoContabilidad=infoFactura.find("obligadoContabilidad").text
            except:
                pass
                
            tipoIdentificacionComprador=infoFactura.find("tipoIdentificacionComprador").text
            razonSocialComprador=infoFactura.find("razonSocialComprador").text
            identificacionComprador=infoFactura.find("identificacionComprador").text
            periodoFiscal=""
            totalSinImpuestos=infoFactura.find("totalSinImpuestos").text
            descuento=""
            try:
                descuento=infoFactura.find("totalDescuento").text
            except:
                pass
            propina=''
            try:
                propina==infoFactura.find("propina").text
            except:
                pass

            importeTotal=""

            try:
                importeTotal=infoFactura.find("importeTotal").text
            except:
                pass
                
            tipoProveedor=""
            if(contribuyenteEspecial==""):
                tipoProveedor="01"
            else:
                tipoProveedor="02"
            
            
            idSociedadInsertar=1
            cur.execute("update documentos set tipoDocumento='"+tipoDoc+"', tipoProveedor='"+tipoProveedor+"', tipoEmision='"+ tipoEmision+"',ambiente='"+ambiente+"',rucProveedor='"+ruc+"',nombreProveedor='"+nombreComercial+"',direccionMatriz='"+dirMatriz+"',direccionEstablecimiento='"+dirEstablecimiento+"',razonSocialP='"+razonSocial+"',tipoIdentificacionC='"+tipoIdentificacionComprador+"',rucCliente='"+identificacionComprador+"',nombreCliente='"+razonSocialComprador+"',establecimiento='"+estab+"',puntoEmision='"+ptoEmi+"',secuencial='"+secuencial+"',fecha='"+fechaEmision+"',totalSinImp='"+totalSinImpuestos+"',descuento='"+descuento+"',propina='"+propina+"',devolucionIva='',valorTotal='"+importeTotal+"',contrEspecial='"+contribuyenteEspecial+"',obligado='"+obligadoContabilidad+"',microempresa='',rimpe='',agenteRetencion='',ejercicioFiscal='' ,idSociedad="+str(idSociedadInsertar)+" WHERE claveAcceso=?", 
            (claveAcceso,))
            conexion.commit()
            for child in impuestos.findall("totalImpuesto"):
                codigo=child.find("codigo").text
                codigoPorcentaje=child.find("codigoPorcentaje").text
                baseImponibleImp=child.find("baseImponible").text
                valor=child.find("valor").text
                    
                conexion.execute("insert into detalleRetFactura (codigo,codigoPorcentaje,baseImponible,valor,idComprobante) values (?,?,?,?,?)",
                (codigo,codigoPorcentaje,baseImponibleImp,valor,idDocumento))
                conexion.commit()
            
            cur.execute("UPDATE documentos set observacion='OK' WHERE id=?",(idDocumento,))
            conexion.commit()

        if(tipoDoc=="N. CRÉDITO"):
            infoNotaCredito=root_comprobante.find("infoNotaCredito")
            #infoTributaria=root_comprobante.find("infoTributaria")
            impuestos=infoNotaCredito.find("totalConImpuestos")
            ambiente=infoTributaria.find("ambiente").text
            tipoEmision=infoTributaria.find("tipoEmision").text
            razonSocial=infoTributaria.find("razonSocial").text
            nombreComercial=""
            try:
                nombreComercial=infoTributaria.find("nombreComercial").text
            except:
                pass
            ruc=infoTributaria.find("ruc").text
            #claveAcceso=infoTributaria.find("claveAcceso").text
            codDoc=infoTributaria.find("codDoc").text
            estab=infoTributaria.find("estab").text
            ptoEmi=infoTributaria.find("ptoEmi").text
            secuencial=infoTributaria.find("secuencial").text
            dirMatriz=infoTributaria.find("dirMatriz").text
            #agenteRetencion=infoTributaria.find("agenteRetencion").text
            
            fechaEmision=infoNotaCredito.find("fechaEmision").text
            dirEstablecimiento=""
            try:
                dirEstablecimiento=infoNotaCredito.find("dirEstablecimiento").text
            except:
                pass
                
            contribuyenteEspecial=""
            obligadoContabilidad=""
            try:
                contribuyenteEspecial=infoNotaCredito.find("contribuyenteEspecial").text
            except:
                pass
            try:
                obligadoContabilidad=infoNotaCredito.find("obligadoContabilidad").text
            except:
                pass
                
            tipoIdentificacionComprador=infoNotaCredito.find("tipoIdentificacionComprador").text
            razonSocialComprador=infoNotaCredito.find("razonSocialComprador").text
            identificacionComprador=infoNotaCredito.find("identificacionComprador").text
            periodoFiscal=""
            totalSinImpuestos=infoNotaCredito.find("totalSinImpuestos").text
            descuento=""
            try:
                descuento=infoNotaCredito.find("totalDescuento").text
            except:
                pass
            propina=''
            try:
                propina==infoNotaCredito.find("propina").text
            except:
                pass

            importeTotal=""

            try:
                importeTotal=infoNotaCredito.find("valorModificacion").text
            except:
                pass
            
            tipoProveedor=""
            if(contribuyenteEspecial==""):
                tipoProveedor="01"
            else:
                tipoProveedor="02"
            
            
            idSociedadInsertar=1

            cur.execute("update documentos set tipoDocumento='"+tipoDoc+"', tipoProveedor='"+tipoProveedor+"', tipoEmision='"+ tipoEmision+"',ambiente='"+ambiente+"',rucProveedor='"+ruc+"',nombreProveedor='"+nombreComercial+"',direccionMatriz='"+dirMatriz+"',direccionEstablecimiento='"+dirEstablecimiento+"',razonSocialP='"+razonSocial+"',tipoIdentificacionC='"+tipoIdentificacionComprador+"',rucCliente='"+identificacionComprador+"',nombreCliente='"+razonSocialComprador+"',establecimiento='"+estab+"',puntoEmision='"+ptoEmi+"',secuencial='"+secuencial+"',fecha='"+fechaEmision+"',totalSinImp='"+totalSinImpuestos+"',descuento='"+descuento+"',propina='"+propina+"',devolucionIva='',valorTotal='"+importeTotal+"',contrEspecial='"+contribuyenteEspecial+"',obligado='"+obligadoContabilidad+"',microempresa='',rimpe='',agenteRetencion='',ejercicioFiscal='' ,idSociedad="+str(idSociedadInsertar)+" WHERE claveAcceso=?", 
            (claveAcceso,))
            conexion.commit()
            codDocModificado=""
            numDocModificado=""
            fechaEmisionDocSustento=""
            try:
                codDocModificado=infoNotaCredito.find("codDocModificado").text
            except:
                pass
            try:
                numDocModificado=infoNotaCredito.find("numDocModificado").text
            except:
                pass
            try:
                fechaEmisionDocSustento=infoNotaCredito.find("fechaEmisionDocSustento").text
            except:
                pass


            for child in impuestos.findall("totalImpuesto"):
                codigo=child.find("codigo").text
                codigoPorcentaje=child.find("codigoPorcentaje").text
                baseImponibleImp=child.find("baseImponible").text
                valor=child.find("valor").text
                    
                conexion.execute("insert into detalleNotaCredito (codDocModificado,numDocModificado,fechaEmisionDocSustento,codigo,codigoPorcentaje,baseImponible,valor,idComprobante) values (?,?,?,?,?,?,?,?)",
                (codDocModificado,numDocModificado,fechaEmisionDocSustento,codigo,codigoPorcentaje,baseImponibleImp,valor,idDocumento))
                conexion.commit()
                
            cur.execute("UPDATE documentos set observacion='OK' WHERE id=?",(idDocumento,))
            conexion.commit()
        if(tipoDoc=="N. DÉBITO"):
            infoNotaDebito=root_comprobante.find("infoNotaDebito")
            #infoTributaria=root_comprobante.find("infoTributaria")
            impuestos=infoNotaDebito.find("impuestos")
            ambiente=infoTributaria.find("ambiente").text
            tipoEmision=infoTributaria.find("tipoEmision").text
            razonSocial=infoTributaria.find("razonSocial").text
            nombreComercial=""
            try:
                nombreComercial=infoTributaria.find("nombreComercial").text
            except:
                pass
            ruc=infoTributaria.find("ruc").text
            
            codDoc=infoTributaria.find("codDoc").text
            estab=infoTributaria.find("estab").text
            ptoEmi=infoTributaria.find("ptoEmi").text
            secuencial=infoTributaria.find("secuencial").text
            dirMatriz=infoTributaria.find("dirMatriz").text
            #agenteRetencion=infoTributaria.find("agenteRetencion").text
            
            fechaEmision=infoNotaDebito.find("fechaEmision").text
            dirEstablecimiento=""
            try:
                dirEstablecimiento=infoNotaDebito.find("dirEstablecimiento").text
            except:
                pass
                
            contribuyenteEspecial=""
            obligadoContabilidad=""
            try:
                contribuyenteEspecial=infoNotaDebito.find("contribuyenteEspecial").text
            except:
                pass
            try:
                obligadoContabilidad=infoNotaDebito.find("obligadoContabilidad").text
            except:
                pass
                
            tipoIdentificacionComprador=infoNotaDebito.find("tipoIdentificacionComprador").text
            razonSocialComprador=infoNotaDebito.find("razonSocialComprador").text
            identificacionComprador=infoNotaDebito.find("identificacionComprador").text
            periodoFiscal=""
            totalSinImpuestos=infoNotaDebito.find("totalSinImpuestos").text
            descuento=""
            try:
                descuento=infoNotaDebito.find("totalDescuento").text
            except:
                pass
            propina=''
            try:
                propina==infoNotaDebito.find("propina").text
            except:
                pass

            importeTotal=""

            try:
                importeTotal=infoNotaDebito.find("valorTotal").text
            except:
                pass
            
            tipoProveedor=""
            if(contribuyenteEspecial==""):
                tipoProveedor="01"
            else:
                tipoProveedor="02"


            idSociedadInsertar=1
            cur.execute("update documentos set tipoDocumento='"+tipoDoc+"', tipoProveedor='"+tipoProveedor+"', tipoEmision='"+ tipoEmision+"',ambiente='"+ambiente+"',rucProveedor='"+ruc+"',nombreProveedor='"+nombreComercial+"',direccionMatriz='"+dirMatriz+"',direccionEstablecimiento='"+dirEstablecimiento+"',razonSocialP='"+razonSocial+"',tipoIdentificacionC='"+tipoIdentificacionComprador+"',rucCliente='"+identificacionComprador+"',nombreCliente='"+razonSocialComprador+"',establecimiento='"+estab+"',puntoEmision='"+ptoEmi+"',secuencial='"+secuencial+"',fecha='"+fechaEmision+"',totalSinImp='"+totalSinImpuestos+"',descuento='"+descuento+"',propina='"+propina+"',devolucionIva='',valorTotal='"+importeTotal+"',contrEspecial='"+contribuyenteEspecial+"',obligado='"+obligadoContabilidad+"',microempresa='',rimpe='',agenteRetencion='',ejercicioFiscal='' ,idSociedad="+str(idSociedadInsertar)+" WHERE claveAcceso=?", 
            (claveAcceso,))
            conexion.commit()
            codDocModificado=""
            numDocModificado=""
            fechaEmisionDocSustento=""
            try:
                codDocModificado=infoNotaDebito.find("codDocModificado").text
            except:
                pass
            try:
                numDocModificado=infoNotaDebito.find("numDocModificado").text
            except:
                pass
            try:
                fechaEmisionDocSustento=infoNotaDebito.find("fechaEmisionDocSustento").text
            except:
                pass


            for child in impuestos.findall("impuesto"):
                codigo=child.find("codigo").text
                codigoPorcentaje=child.find("codigoPorcentaje").text
                baseImponibleImp=child.find("baseImponible").text
                valor=child.find("valor").text
                    
                conexion.execute("insert into detalleNotaDebito (codDocModificado,numDocModificado,fechaEmisionDocSustento,codigo,codigoPorcentaje,baseImponible,valor,idComprobante) values (?,?,?,?,?,?,?,?)",
                (codDocModificado,numDocModificado,fechaEmisionDocSustento,codigo,codigoPorcentaje,baseImponibleImp,valor,idDocumento))
                conexion.commit()
        
            cur.execute("UPDATE documentos set observacion='OK' WHERE id=?",(idDocumento,))
            conexion.commit()
path = rutaArchivos
dir_list = os.listdir(path)

db="BDSRI.s3db"
conexion1=sqlite3.connect(db)
cur = conexion1.cursor()
current_date = dt.datetime.now()
fechaActual=current_date.strftime("%d/%m/%Y")#Path 
insert=cur.execute("insert into ejecucion(fecha) values (?)",(fechaActual,))
conexion1.commit()
idEjec=insert.lastrowid
conexion1.close()
for file in dir_list:
    leerXML(rutaArchivos+"\\"+file,db,idEjec)
generaXls(db,idEjec)
